#! python3
# daily_osd.py

#a program to prepare the live osd> Rs 100 from zdisfolup_alv
from pandas import DataFrame , Series
import pandas as pd
import numpy as np
import openpyxl as xl, time, os
#from openpyxl.styles import Color, PatternFill, Font, Border
from datetime import date , datetime

def clearFile(filename):
    '''
    get a file and delete all data in that file
    argument: a filename with extension
    return: None
    '''
    fileObj= xl.load_workbook(filename)
    sheet= fileObj.active
    max_row, max_col= sheet.max_row, sheet.max_column
    for i in range(2, max_row+1):
        for j in range(1, max_col+1):
            sheet.cell(row= i, column= j).value= None     #assignin None value to a cell means clearing the cell
    print(filename, " --> content cleared")
    fileObj.save(filename)

def findCell(sheet, subject):
    '''
    get a sheet object and a subject(string) and return the cell index of the first occurance
    of that subject in that sheet
    argument= sheet object, string
    return= integer, integer
    '''
    max_row, max_col= sheet.max_row, sheet.max_column
    for i in range(1, max_row+1):
        for j in range(1, max_col+1):
            if str(sheet.cell(row= i, column= j).value).strip()== subject.strip():  #strip() is necessary in SAP
                return i, j
            
def osd(fileName , pay_data):
    CCC= ['103.xlsx', '201.xlsx', '202.xlsx', '208.xlsx', '300.xlsx', '301.xlsx','401.xlsx','402.xlsx']
    fileObj= xl.load_workbook(fileName)
    newSheet= fileObj.active
    newFileRowCount= 2  #to store current row number in output files
##    input("What should we consider as current date: ? ")
##    _year= int(input("enter year "))
##    _month = int(input("enter month "))
##    _date = int(input("enter date "))
    _year , _month , _day = str(datetime.today().replace(day = 1))[:10].split('-')
    _year , _month , _day = int(_year) , int(_month) , int(_day)
    todays_date = date(_year , _month , _day)
    
    for eachFileName in CCC:
        print("data preparation going on for file: ", eachFileName)
        eachFileObj= xl.load_workbook(eachFileName)
        eachSheet= eachFileObj.active
        
        #starting of block which captures the cell indices of each heading. 
        mruRow, mruCol= findCell(eachSheet, 'MRU')
        conidRow, conidCol= findCell(eachSheet, 'Consumer Id')
        nameRow, nameCol= findCell(eachSheet, 'Name')
        addressRow, addressCol= findCell(eachSheet, 'Address')
        bclassRow, bclassCol= findCell(eachSheet, 'Base Class')
        deviceRow, deviceCol= findCell(eachSheet, 'Device')
        periodRow, periodCol= findCell(eachSheet, 'O/S Duedate Range')
        osdRow, osdCol= findCell(eachSheet, 'D2 Net O/S')   #for checking osd amount
        mobRow, mobCol= findCell(eachSheet, 'Mobile Number')
        disRow, disCol= findCell(eachSheet, 'Discon Status')   #for checking live
        disDateRow, disDateCol= findCell(eachSheet, 'Discon Date')   #for checking live
        depRow, depCol= findCell(eachSheet, 'Nature of Conn')  #for checking govt / non-govt
        govRow, govCol= findCell(eachSheet, "Gov/Non-Gov")
        latRow, latCol= findCell(eachSheet, "Latitude")
        longRow, longCol= findCell(eachSheet, "Longitude")
        #end of the block which captures the cell indices of each heading. 

        #here comes the list of governement department. include any new government department name
        #as per the zdisfolup_alv
        '''
        govt_class= ['AGRI IRRI.','JUDICIAL', 'GRAM PANCHAYET', 'Municipality', 'Nature of Conn.', 'PHE','IRRIG. AND WATER WAYS',
                'POLICE', 'POST AND TELEGRAPH', 'PWD(ELECTRICAL)', 'STATE GOVT.','AGRI MECH','MINOR IRRIG.',
                 'DEPT. OF HEALTH','PRI HEALTH CENTRE','MDTW','ZILLA PARISAD','PRIMARY SCHOOL','B.D.O. OFFICE',
                 'PWD(CONST. BOARD)','DEPT. OF EDU.','C.A.D.A.','JUDICIAL','PWD','PWD(NH DIVISION)','HIGH SCHOOL',
                 'PWD(ROADS)','DEPT. OF AGRI.']
        '''
        govt_class= []
        '''
        redFill = PatternFill(start_color='FFA07A',
                   end_color='FFA07A',
                   fill_type='solid')
        '''
        for i in range(1, eachSheet.max_row+1):
            #print(i, eachSheet.cell(row= i, column= osdCol).value)
            try:
                osdAmount= float(eachSheet.cell(row= i, column= osdCol).value)
            except:
                continue
            mru= eachSheet.cell(row= i, column= mruCol).value
            conid = eachSheet.cell(row= i, column= conidCol).value
            osdRemark= "N/A"
            if osdAmount>= 200:  #if osd is more than given amount
                if osdAmount> 10000: osdRemark= "1. OSD> Rs. 10000"
                elif osdAmount> 5000: osdRemark= "2. Rs. 10000> OSD> Rs. 5000"
                elif osdAmount> 3000: osdRemark= "3. Rs. 5000> OSD> 3000"
                elif osdAmount> 1000: osdRemark= "4. Rs. 3000> OSD> Rs. 1000"
                elif osdAmount> 500: osdRemark= "5. Rs. 1000> OSD> Rs. 500"
                else: osdRemark= "6. OSD< Rs. 500"
                #print(eachFileName, newFileRowCount)
                #print(mru,  type(mru), mru[:2], type(mru[:2]))
                if mru is not None:
                    if mru[:2]== '7D': cccCode= '3332103'
                    elif mru[:2]== 'GD': cccCode= '3332201'
                    elif mru[:2]== 'C1': cccCode= '3332202'
                    elif mru[:2]== 'IA': cccCode= '3332208'
                    elif mru[:2]== '7C': cccCode= '3332300'
                    elif mru[:2]== '8O': cccCode= '3332301'
                    elif mru[:2]== 'AU': cccCode= '3332401'
                    else: cccCode= '3332402'
                else:
                    continue
                #print(eachFileName, newFileRowCount)
                newSheet.cell(row= newFileRowCount, column= 1).value= newFileRowCount- 1
                newSheet.cell(row= newFileRowCount, column= 2).value= cccCode
                newSheet.cell(row= newFileRowCount, column= 3).value= eachSheet.cell(row= i, column= mruCol).value
                newSheet.cell(row= newFileRowCount, column= 4).value= conid
                newSheet.cell(row= newFileRowCount, column= 5).value= osdRemark
                #print(eachFileName, newFileRowCount)
                ######newSheet.cell(row= newFileRowCount, column= 5).fill= redFill
                #print(eachFileName, newFileRowCount)
                newSheet.cell(row= newFileRowCount, column= 6).value= eachSheet.cell(row= i, column= nameCol).value
                newSheet.cell(row= newFileRowCount, column= 7).value= eachSheet.cell(row= i, column= addressCol).value
                newSheet.cell(row= newFileRowCount, column= 8).value= eachSheet.cell(row= i, column= bclassCol).value
                newSheet.cell(row= newFileRowCount, column= 9).value= eachSheet.cell(row= i, column= deviceCol).value
                newSheet.cell(row= newFileRowCount, column= 10).value= eachSheet.cell(row= i, column= periodCol).value
                newSheet.cell(row= newFileRowCount, column= 11).value= eachSheet.cell(row= i, column= osdCol).value
                #print(eachFileName, newFileRowCount)
                newSheet.cell(row= newFileRowCount, column= 12).value= eachSheet.cell(row= i, column= osdCol).value/100000
                
                if str(conid).strip() in pay_data:
                    pay_date = pay_data[str(conid)]
                    #print(pay_date ,int(pay_date[:4]) , int(pay_date[4:6]), int(pay_date[6:])) 
                    pay_date = date(int(pay_date[:4]) , int(pay_date[4:6]), int(pay_date[6:]))
                else:
                    pay_date = date(2000 , 1 , 1)
                #pay_date = str(pay_date)
                if osdAmount > 5000:
                    newSheet.cell(row= newFileRowCount, column= 13).value= "OSD > 5k"
                else:
                    newSheet.cell(row= newFileRowCount, column= 13).value= "OSD < 5K"
                newSheet.cell(row= newFileRowCount, column= 14).value= pay_date
                days_till_last_paid = (todays_date - pay_date).days
                newSheet.cell(row= newFileRowCount, column= 15).value= days_till_last_paid
                if days_till_last_paid > 365:
                    pay_remark = "1 Year"
                elif days_till_last_paid > 300 :
                    pay_remark = "10 Months"
                elif days_till_last_paid > 180 :
                    pay_remark = "6 Months"
                elif days_till_last_paid > 120 :
                    pay_remark = "4 Months"
                elif days_till_last_paid > 90 :
                    pay_remark = "3 Months"
                elif days_till_last_paid > 30:
                    pay_remark = "1 Month"
                else:
                    pay_remark = "Within 1 Month"
                newSheet.cell(row= newFileRowCount, column= 16).value= pay_remark
                newSheet.cell(row= newFileRowCount, column= 17).value= eachSheet.cell(row= i, column= mobCol).value
                #print(eachFileName, newFileRowCount)
                newSheet.cell(row= newFileRowCount, column= 18).value= eachSheet.cell(row= i, column= disCol).value
                #print(eachFileName, newFileRowCount)
                newSheet.cell(row= newFileRowCount, column= 19).value= eachSheet.cell(row= i, column= disDateCol).value
                #print(eachFileName, newFileRowCount)
                newSheet.cell(row= newFileRowCount, column= 20).value= eachSheet.cell(row= i, column= govCol).value
                newSheet.cell(row= newFileRowCount, column= 21).value= eachSheet.cell(row= i, column= depCol).value
                newSheet.cell(row= newFileRowCount, column= 22).value= eachSheet.cell(row= i, column= latCol).value
                newSheet.cell(row= newFileRowCount, column= 23).value= eachSheet.cell(row= i, column= longCol).value
                newFileRowCount+= 1

        print("Data found : ", newFileRowCount)            
    newSheet.freeze_panes= "F2"                    
    fileObj.save(fileName)
    print('Output file is ', fileName)

def all_pay(filename):
    fileObj= xl.load_workbook(filename)
    sheet = fileObj.active
    pay_data = {}

    for i in range(1, sheet.max_row + 1):
        con_id= str(sheet.cell(row = i , column = 1).value).strip()
        pay_date= str(sheet.cell(row = i , column = 2).value).strip()
        pay_data[con_id.strip()]= pay_date.strip()

    return pay_data

def dm_format_osd(filename):
    print("\nNow DM / RM format OSD file will be created: \n")
    ccc_list = [3332103 , 3332201 , 3332202 , 3332208 , 3332300 , 3332301 , 3332401 , 3332402]
    columns = ['CCC Code', 'MRU', 'Con Id', 'Name', 'Address','Base Class', 'Device', 'Period', 'OSD', 'OSD Lakh', 'MOBILE NO']

    data = pd.read_excel(filename)

    live_non_govt_dom_condition = (data['Base Class'] == 'D') & (data['Discon Status'].isna()) & (data['Govt Status'].isna())
    live_non_govt_comm_condition = (data['Base Class'] == 'C') & (data['Discon Status'].isna()) & (data['Govt Status'].isna())
    live_non_govt_ind_condition = (data['Base Class'] == 'I') & (data['Discon Status'].isna()) & (data['Govt Status'].isna())
    live_non_govt_agri_condition = (data['Base Class'] == 'A') & (data['Discon Status'].isna()) & (data['Govt Status'].isna())
    live_non_govt_oth_condition = (~data['Base Class'].isin(['D' , 'C' , 'I' , 'A'])) & (data['Discon Status'].isna()) & (data['Govt Status'].isna())

    dom_osd = data[live_non_govt_dom_condition]
    comm_osd = data[live_non_govt_comm_condition]
    ind_osd = data[live_non_govt_ind_condition]
    agri_osd = data[live_non_govt_agri_condition]
    oth_osd = data[live_non_govt_oth_condition]

    dom_osd_10000 = dom_osd[dom_osd['OSD'] > 10000]
    dom_osd_5000 = dom_osd[dom_osd['OSD'] > 5000]
    dom_osd_3000 = dom_osd[dom_osd['OSD'] > 3000]
    dom_osd_1000 = dom_osd[dom_osd['OSD'] > 1000]
    dom_osd_500 = dom_osd[dom_osd['OSD'] > 500]
    dom_osd_5000_to_10000 = dom_osd[dom_osd['OSD'].between(5000 , 9999)]
    dom_osd_3000_to_5000 = dom_osd[dom_osd['OSD'].between(3000 , 4999)]
    dom_osd_1000_to_3000 = dom_osd[dom_osd['OSD'].between(1000 , 3000)]
    dom_osd_500_to_1000 = dom_osd[dom_osd['OSD'].between(500 , 1000)]

    comm_osd_500 = comm_osd[comm_osd['OSD'] > 500]
    comm_osd_3000 = comm_osd[comm_osd['OSD'] > 3000]
    comm_osd_5000 = comm_osd[comm_osd['OSD'] > 5000]

    ind_osd_summary = ind_osd.groupby('CCC Code').agg({'Con Id' : 'count' , 'OSD Lakh' : 'sum'}).reindex(ccc_list , fill_value = 0)
    ind_osd_summary = pd.concat([ind_osd_summary , ind_osd_summary.sum().to_frame().T])
    ind_osd_summary.rename(index = {0 : 'TOTAL'} , inplace = True)

    dom_osd_summary = dom_osd.groupby('CCC Code').agg({'Con Id' : 'count' , 'OSD Lakh' : 'sum'}).reindex(ccc_list , fill_value = 0)
    dom_osd_summary = pd.concat([dom_osd_summary , dom_osd_summary.sum().to_frame().T])
    dom_osd_summary.rename(index = {0 : 'TOTAL'} , inplace = True)

    dom_osd_10000_summary = dom_osd_10000.groupby('CCC Code').agg({'Con Id' : 'count' , 'OSD Lakh' : 'sum'}).reindex(ccc_list , fill_value = 0)
    dom_osd_10000_summary = pd.concat([dom_osd_10000_summary , dom_osd_10000_summary.sum().to_frame().T])
    dom_osd_10000_summary.rename(index = {0 : 'TOTAL'} , inplace = True)

    dom_osd_5000_to_10000_summary = dom_osd_5000_to_10000.groupby('CCC Code').agg({'Con Id' : 'count' , 'OSD Lakh' : 'sum'}).reindex(ccc_list , fill_value = 0)
    dom_osd_5000_to_10000_summary = pd.concat([dom_osd_5000_to_10000_summary , dom_osd_5000_to_10000_summary.sum().to_frame().T])
    dom_osd_5000_to_10000_summary.rename(index = {0 : 'TOTAL'} , inplace = True)

    dom_osd_3000_to_5000_summary = dom_osd_3000_to_5000.groupby('CCC Code').agg({'Con Id' : 'count' , 'OSD Lakh' : 'sum'}).reindex(ccc_list , fill_value = 0)
    dom_osd_3000_to_5000_summary = pd.concat([dom_osd_3000_to_5000_summary , dom_osd_3000_to_5000_summary.sum().to_frame().T])
    dom_osd_3000_to_5000_summary.rename(index = {0 : 'TOTAL'} , inplace = True)

    dom_osd_3000_summary = dom_osd_3000.groupby('CCC Code').agg({'Con Id' : 'count' , 'OSD Lakh' : 'sum'}).reindex(ccc_list , fill_value = 0)
    dom_osd_3000_summary = pd.concat([dom_osd_3000_summary , dom_osd_3000_summary.sum().to_frame().T])
    dom_osd_3000_summary.rename(index = {0 : 'TOTAL'} , inplace = True)

    # dom_osd_3000_to_5000_summary = dom_osd_3000_to_5000.groupby('CCC Code').agg({'Con Id' : 'count' , 'OSD Lakh' : 'sum'}).reindex(ccc_list , fill_value = 0)
    # dom_osd_3000_to_5000_summary = pd.concat([dom_osd_3000_to_5000_summary , dom_osd_3000_to_5000_summary.sum().to_frame().T])
    # dom_osd_3000_to_5000_summary.rename(index = {0 : 'TOTAL'} , inplace = True)

    comm_osd_summary = comm_osd.groupby('CCC Code').agg({'Con Id' : 'count' , 'OSD Lakh' : 'sum'}).reindex(ccc_list , fill_value = 0)
    comm_osd_summary = pd.concat([comm_osd_summary , comm_osd_summary.sum().to_frame().T])
    comm_osd_summary.rename(index = {0 : 'TOTAL'} , inplace = True)

    comm_osd_500_summary = comm_osd_500.groupby('CCC Code').agg({'Con Id' : 'count' , 'OSD Lakh' : 'sum'}).reindex(ccc_list , fill_value = 0)
    comm_osd_500_summary = pd.concat([comm_osd_500_summary , comm_osd_500_summary.sum().to_frame().T])
    comm_osd_500_summary.rename(index = {0 : 'TOTAL'} , inplace = True)

    comm_osd_5000_summary = comm_osd_5000.groupby('CCC Code').agg({'Con Id' : 'count' , 'OSD Lakh' : 'sum'}).reindex(ccc_list , fill_value = 0)
    comm_osd_5000_summary = pd.concat([comm_osd_5000_summary , comm_osd_5000_summary.sum().to_frame().T])
    comm_osd_5000_summary.rename(index = {0 : 'TOTAL'} , inplace = True)

    agri_osd_summary = agri_osd.groupby('CCC Code').agg({'Con Id' : 'count' , 'OSD Lakh' : 'sum'}).reindex(ccc_list , fill_value = 0)
    agri_osd_summary = pd.concat([agri_osd_summary , agri_osd_summary.sum().to_frame().T])
    agri_osd_summary.rename(index = {0 : 'TOTAL'} , inplace = True)

    oth_osd_summary = oth_osd.groupby('CCC Code').agg({'Con Id' : 'count' , 'OSD Lakh' : 'sum'}).reindex(ccc_list , fill_value = 0)
    oth_osd_summary = pd.concat([oth_osd_summary , oth_osd_summary.sum().to_frame().T])
    oth_osd_summary.rename(index = {0 : 'TOTAL'} , inplace = True)

    def top_100(non_govt_live_df):
        top_100_df = []
        for each_ccc in ccc_list:
            sorted_temp = non_govt_live_df[ (non_govt_live_df['CCC Code'] == int(each_ccc) )]
            sorted_temp = sorted_temp.sort_values(by = 'OSD' , ascending = False)[:100]
            top_100_df.append(sorted_temp)

        top_100_df = pd.concat(top_100_df)
        return top_100_df

    top_100_dom = top_100(dom_osd)
    top_100_dom_summary = top_100_dom.groupby('CCC Code').agg({'Con Id' : 'count' , 'OSD Lakh' : 'sum'}).reindex(ccc_list , fill_value = 0)
    top_100_dom_summary = pd.concat([top_100_dom_summary , top_100_dom_summary.sum().to_frame().T])
    top_100_dom_summary.rename(index = {0 : 'TOTAL'} , inplace = True)

    top_100_comm = top_100(comm_osd)
    top_100_comm_summary = top_100_comm.groupby('CCC Code').agg({'Con Id' : 'count' , 'OSD Lakh' : 'sum'}).reindex(ccc_list , fill_value = 0)
    top_100_comm_summary = pd.concat([top_100_comm_summary , top_100_comm_summary.sum().to_frame().T])
    top_100_comm_summary.rename(index = {0 : 'TOTAL'} , inplace = True)

    output_file = 'OSD-SUMMARY-' + str(datetime.now())[:-7].replace(":","-").replace(" ","-") + '.xlsx'
    with pd.ExcelWriter(output_file , engine="xlsxwriter") as writer:
     
        dom_osd_summary.to_excel(writer , sheet_name = 'summary' , startrow = 2 , startcol = 0)
        worksheet = writer.sheets['summary']
        worksheet.write(0 , 1 , "DOM-ALL-OSD")

        dom_osd_10000_summary.to_excel(writer , sheet_name = 'summary' , startrow = 2 , startcol = 4)
        worksheet.write(0 , 5 , "DOM-ABOVE-10000")

        dom_osd_5000_to_10000_summary.to_excel(writer , sheet_name = 'summary' , startrow = 2 , startcol = 8)
        worksheet.write(0 , 9 , "DOM-5000-TO-10000")

        dom_osd_3000_to_5000_summary.to_excel(writer , sheet_name = 'summary' , startrow = 2 , startcol = 12)
        worksheet.write(0 , 13 , "DOM-3000-TO-5000")

        top_100_dom_summary.to_excel(writer , sheet_name = 'summary' , startrow = 2 , startcol = 16)
        worksheet.write(0 , 17 , "DOM-TOP-100-OSD")

        dom_osd_3000_summary.to_excel(writer , sheet_name = 'summary' , startrow = 2 , startcol = 20)
        worksheet.write(0 , 21 , "DOM-ABOVE-3000")

        comm_osd_summary.to_excel(writer , sheet_name = 'summary' , startrow = 15 , startcol = 0)
        worksheet.write(13 , 1 , "COMM-ALL-OSD")

        comm_osd_5000_summary.to_excel(writer , sheet_name = 'summary' , startrow = 15 , startcol = 4)
        worksheet.write(13 , 5 , "COMM-ABOVE-5000")

        comm_osd_500_summary.to_excel(writer , sheet_name = 'summary' , startrow = 15 , startcol = 8)
        worksheet.write(13 , 9 , "COMM-ABOVE-500")

        top_100_comm_summary.to_excel(writer , sheet_name = 'summary' , startrow = 15 , startcol = 12)
        worksheet.write(13 , 13 , "COMM-TOP-100-OSD")

        ind_osd_summary.to_excel(writer , sheet_name = 'summary' , startrow = 28 , startcol = 0)
        worksheet.write(26 , 1 , "IND-ALL-OSD")


        dom_osd_10000[columns].to_excel(writer, sheet_name= 'dom_10K', startrow= 0 , index = False)
        dom_osd_5000_to_10000[columns].to_excel(writer, sheet_name= 'dom_5K_to_10K', startrow= 0 , index = False)
        dom_osd_3000_to_5000[columns].to_excel(writer, sheet_name= 'dom_3K_to_5K', startrow= 0 , index = False)
        top_100_dom[columns].to_excel(writer, sheet_name= 'dom_top_100', startrow= 0 , index = False)
        comm_osd_5000[columns].to_excel(writer, sheet_name= 'comm_5000', startrow= 0 , index = False)
        comm_osd_500[columns].to_excel(writer, sheet_name= 'comm_500', startrow= 0 , index = False)
        top_100_comm[columns].to_excel(writer, sheet_name= 'comm_top_100', startrow= 0 , index = False)
        ind_osd[columns].to_excel(writer, sheet_name= 'ind_osd_all', startrow= 0 , index = False)

    
if __name__== '__main__':
    start= time.time()
    fileName= "kalyani_total_osd.xlsx"
    clearFile(fileName)
    pay_data= all_pay('all_pay.xlsx')
    osd(fileName , pay_data)
    dm_format_osd(fileName)
    end= time.time()
    print("Finished in %.2f minutes. Please check your file" %float((end-start)/60))
    input("Press ENTER")
